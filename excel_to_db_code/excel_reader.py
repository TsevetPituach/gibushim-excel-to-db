from typing import Any, Dict, Generator, Tuple
from openpyxl.reader.excel import load_workbook

from .validators import is_half_empty


HalfDict = Dict[str, Any]


def _slice_to_half(values: Tuple[Any, ...]) -> HalfDict:
    keys = ("group_id", "chest_number", "candidate_name", "assessor_name", "grade", "comment")
    return {k: (values[i] if len(values) > i else None) for i, k in enumerate(keys)}


def _row_to_halves(row_values: Tuple[Any, ...], half2_offset: int = 15) -> Tuple[HalfDict, HalfDict]:
    # Columns A-F (0..5), O/P..T/U (half2_offset..half2_offset+5)
    # Both halves map to the same schema keys.
    a_f = row_values[0:6] if len(row_values) >= 6 else tuple([None] * 6)
    end = half2_offset + 6
    second = row_values[half2_offset:end] if len(row_values) >= end else tuple([None] * 6)

    half1 = _slice_to_half(a_f)
    half2 = _slice_to_half(second)

    return half1, half2

def iter_excel_halves(
    path: str,
) -> Generator[Tuple[int, int, HalfDict], None, None]:
    """Yield (row_number, half_index, half_dict) for each non-empty half in Excel.
    half_index is 1 for A-F and 2 for P-U.
    Row numbers are 1-based as in Excel.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    start = 2 # because it has header if not it will be 1
    for i, row in enumerate(ws.iter_rows(min_row=start, values_only=True), start=start):
        half1, half2 = _row_to_halves(row)

        if not is_half_empty(half1):
            yield (i, 1, half1)
        if not is_half_empty(half2):
            yield (i, 2, half2)
